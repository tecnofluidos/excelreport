import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
import yaml

class TExcelReport:
    def __init__(self, sheetname, dataset, yaml_file):
        self.wb = Workbook()
        self.ws = self.wb.active        
        self.ws.title = sheetname  
        self.dataset = dataset
        self.__load_yaml__(yaml_file)
        self.__create_global_styles__()
        self.__apply_cell_styles__()
        self.current_row = 1

    def __del__(self):
        if self.wb is not None:
            self.wb.close()

    # File --------------------------------------------------------------------------------------------------
    def __load_yaml__(self, yaml_file):
        with open(yaml_file, 'r', encoding='utf-8') as file:
            self.config = yaml.safe_load(file)

    def save(self, filename):
        self.wb.save(filename)

    def build(self):
        self.__apply_master_data__()
        
    # Format & Style -----------------------------------------------------------------------------------------
    def __column_resize__(self, column, size):
        resize = (size * 1.05777054515867)
        self.ws.column_dimensions[column].width = resize      

    def setCell(self, cell_coordinate, value, style=None):
        cell = self.ws[cell_coordinate]
        cell.value = value
        if style:
            self.__apply_named_style__(cell, style)

    def __create_global_styles__(self):
        global_config = self.config.get('styles', {})
        for styleDef, style_config in global_config.items():
            style_name = style_config['style'].get('name', f"global_{styleDef}_style")
            self.create_named_style(style_config['style'], style_name)

    def __apply_cell_styles__(self):
        cells = self.config.get('cells', {})
        for cell_coord, cell_config in cells.items():
            style_name = cell_config['style'].get('use', cell_config['style'].get('name', f"cell_{cell_coord}_style"))
            if 'merge' in cell_config:
                self.ws.merge_cells(cell_config['merge'])
                merge_range = cell_config['merge']
                start_cell, end_cell = merge_range.split(':')
                start_row, start_col = openpyxl.utils.cell.coordinate_to_tuple(start_cell)
                end_row, end_col = openpyxl.utils.cell.coordinate_to_tuple(end_cell)
                for row in range(start_row, end_row + 1):
                    for col in range(start_col, end_col + 1):
                        cell = self.ws.cell(row=row, column=col)
                        if style_name in self.wb.named_styles:
                            self.__apply_named_style__(cell, style_name)
            else:
                cell = self.ws[cell_coord]
                if style_name in self.wb.named_styles:
                    self.__apply_named_style__(cell, style_name)

    def __get_named_style_index__(self, style_name):
        for index, ns in enumerate(self.wb.named_styles):
            if ns == style_name:
                return index
        return -1
    
    def create_named_style(self, style_dict, name):        
        index = self.__get_named_style_index__(name)
        if index != -1:
            return self.wb.named_styles[index]   

        named_style = NamedStyle(name=name)
        if 'font' in style_dict:
            font_config = style_dict['font']
            named_style.font = Font(
                name=font_config.get('name', 'Calibri'),
                size=font_config.get('size', 11),
                bold=font_config.get('bold', False),
                italic=font_config.get('italic', False),
                vertAlign=font_config.get('vertAlign', None),
                underline=font_config.get('underline', 'none'),
                strike=font_config.get('strike', False),
                color=font_config.get('color', '000000')
            )
        if 'alignment' in style_dict:
            alignment_config = style_dict['alignment']
            named_style.alignment = Alignment(
                horizontal=alignment_config.get('horizontal', 'general'),
                vertical=alignment_config.get('vertical', 'bottom'),
                text_rotation=alignment_config.get('text_rotation', 0),
                wrap_text=alignment_config.get('wrap_text', False),
                shrink_to_fit=alignment_config.get('shrink_to_fit', False),
                indent=alignment_config.get('indent', 0)
            )
        if 'border' in style_dict:
            border_config = style_dict['border']
            sides = {}
            for side in border_config.get('side', ['top', 'bottom', 'left', 'right']):
                sides[side] = Side(
                    border_style=border_config.get('style', 'thin'),
                    color=border_config.get('color', '000000')
                )
            named_style.border = Border(**sides)
        if 'fill' in style_dict:
            fill_config = style_dict['fill']            
            named_style.fill = PatternFill(
                patternType=fill_config.get('patternType', 'solid'),
                fgColor=fill_config.get('start_color', 'FFFFFF'),
                bgColor=fill_config.get('end_color', 'FFFFFF')
            )
        if 'format' in style_dict:
            named_style.number_format = style_dict.get('format', 'General')
        
        self.wb.add_named_style(named_style)
        return named_style

    def __apply_named_style__(self, cell, style_config):
        if isinstance(style_config, str):
            named_style = self.create_named_style({}, style_config)
        else:
            style_name = style_config.get('use', style_config.get('name', f"cell_{cell.coordinate}_style"))
            named_style = self.create_named_style(style_config, style_name)
        
        cell.style = named_style
        
        if 'format' in style_config:
            cell.number_format = style_config['format']

        if 'font' in style_config:
            font_config = style_config['font']
            cell.font = Font(
                name=font_config.get('name', cell.font.name),
                size=font_config.get('size', cell.font.size),
                bold=font_config.get('bold', cell.font.bold),
                italic=font_config.get('italic', cell.font.italic),
                vertAlign=font_config.get('vertAlign', cell.font.vertAlign),
                underline=font_config.get('underline', cell.font.underline),
                strike=font_config.get('strike', cell.font.strike),
                color=font_config.get('color', cell.font.color)
            )
        if 'alignment' in style_config:
            alignment_config = style_config['alignment']
            cell.alignment = Alignment(
                horizontal=alignment_config.get('horizontal', cell.alignment.horizontal),
                vertical=alignment_config.get('vertical', cell.alignment.vertical),
                text_rotation=alignment_config.get('text_rotation', cell.alignment.text_rotation),
                wrap_text=alignment_config.get('wrap_text', cell.alignment.wrap_text),
                shrink_to_fit=alignment_config.get('shrink_to_fit', cell.alignment.shrink_to_fit),
                indent=alignment_config.get('indent', cell.alignment.indent)
            )
        if 'border' in style_config:
            border_config = style_config['border']
            sides = {}
            for side in border_config.get('side', ['top', 'bottom', 'left', 'right']):
                sides[side] = Side(
                    border_style=border_config.get('style', 'thin'),
                    color=border_config.get('color', '000000')
                )
            cell.border = Border(**sides)
        if 'fill' in style_config:
            fill_config = style_config['fill']
            cell.fill = PatternFill(
                patternType=fill_config.get('patternType', cell.fill.patternType),
                fgColor=fill_config.get('start_color', cell.fill.fgColor),
                bgColor=fill_config.get('end_color', cell.fill.bgColor)
            )

    # Group Header---------------------------------------------------------------------------------------------
    def __apply_group_header__(self):
        group_header_config = self.config.get('dataset', {}).get('break_fields', {})
        start_cell = self.config.get('dataset', {}).get('start', 'A6')
        self.current_row = openpyxl.utils.cell.coordinate_to_tuple(start_cell)[0]

        break_data = group_header_config.get('data', {})
        column_defs = self.config.get('dataset', {}).get('columns', {}).get('columnDefs', {})

        for record in self.dataset:
            for field in group_header_config['fields']:
                if record.get(field) != getattr(self, f"current_{field}", None):
                    self.__print_group_header__(break_data, column_defs, record)
                    for bf in group_header_config['fields']:
                        setattr(self, f"current_{bf}", record.get(bf))

    def __print_group_header__(self, break_data, column_defs, record):
        self.__print_break_fields__(break_data, record)
        self.__print_column_headers__(column_defs)
        
    def __print_break_fields__(self, break_data, record):
        for col, config in break_data.items():
            cell_coord = f"{col}{self.current_row}"
            field = config['field']
            
            self.setCell(cell_coord, record[field], config['style'])
            if 'merge' in config:
                merge_range = config['merge'].replace('%', str(self.current_row))
                self.ws.merge_cells(merge_range)

        self.current_row += 1

    def __print_column_headers__(self, column_defs):
        for col, config in column_defs.items():
            cell_coord = f"{col}{self.current_row}"
            header = config['header']['value']
            size = config.get('width', 8.43)
            self.__column_resize__(col, size)
            self.setCell(cell_coord, header, config['header'].get('style', {}))
        self.current_row += 1

    # Dataset -------------------------------------------------------------------------------------------------
    def __apply_master_data__(self):
        group_header_config = self.config.get('dataset', {}).get('break_fields', {})
        start_cell = self.config.get('dataset', {}).get('start', 'A6')
        self.current_row = openpyxl.utils.cell.coordinate_to_tuple(start_cell)[0]

        break_data = group_header_config.get('data', {})
        column_defs = self.config.get('dataset', {}).get('columns', {}) 
        first_data_row = None  # Guarda a linha do primeiro dado impresso no grupo

        for record in self.dataset:
            # Antes de Aplicar o cabeçalho verificar se precisa imprimir o rodapé do grupo
            if first_data_row is not None:
                for field in group_header_config['fields']:
                    if record.get(field) != getattr(self, f"current_{field}", None):
                        self.__apply_footer__(column_defs, first_data_row, self.current_row - 1)
                        first_data_row = None
                        break

            # Aplica cabeçalho de grupo
            for field in group_header_config['fields']:
                if record.get(field) != getattr(self, f"current_{field}", None):
                    self.__print_group_header__(break_data, column_defs, record)
                    for bf in group_header_config['fields']:
                        setattr(self, f"current_{bf}", record.get(bf))
                    break

            # Aplica os dados do dataset
            if first_data_row is None:
                first_data_row = self.current_row  # Primeira linha de dados do grupo

            for col, config in column_defs.items():
                cell_coord = f"{col}{self.current_row}"
                field = config.get('field', None)
                if field and field in record:
                    self.setCell(cell_coord, record[field], config.get('style', {}))
                else:
                    self.setCell(cell_coord, record[config['data']['field']], config['data'].get('style', {}))
            self.current_row += 1

        # Apenas na última impressão, pois não terá outro grupo
        if first_data_row is not None:
            self.__apply_footer__(column_defs, first_data_row, self.current_row - 1)  # Passa start_row e end_row

    # Group footer---------------------------------------------------------------------------------------------
    def __apply_footer__(self, column_defs, start_row, end_row):
       
        if not column_defs:
            return

        for col, config in column_defs.items():

            formula_map = {
                'SUM': f"=SUM({col}{start_row}:{col}{end_row})",
                'AVG': f"=AVERAGE({col}{start_row}:{col}{end_row})",
                'COUNT': f"=COUNT({col}{start_row}:{col}{end_row})",
                'MIN': f"=MIN({col}{start_row}:{col}{end_row})",
                'MAX': f"=MAX({col}{start_row}:{col}{end_row})"
            }            

            cell_coord = f"{col}{self.current_row}"
            
            footer_config = config.get('footer', {})

            value = footer_config.get('value','')
            formula = footer_config.get('formula','').upper()
            style = footer_config.get('style',{})

            if value:
               self.setCell(cell_coord, value.replace("/*",""), style) 
            elif formula in formula_map:
                value = formula_map[formula]
                self.setCell(cell_coord, value, style)

            if 'merge' in footer_config:
                merge_range = footer_config['merge'].replace('%', str(self.current_row))
                self.ws.merge_cells(merge_range)    
        
        self.current_row += 2
        return


